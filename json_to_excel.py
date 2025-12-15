import json
import pandas as pd
import os

def json_to_excel(json_path, output_path):
    # Load JSON data
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: File not found at {json_path}")
        return
    except Exception as e:
        print(f"Error reading JSON: {e}")
        return

    # Prepare data for DataFrame
    rows = []
    
    # Check if data has new structure with 'summary' and 'data'
    if "data" in data and "summary" in data:
        items_dict = data["data"]
        summary = data["summary"]
        total_time = summary.get("total_process_time", 0)
    else:
        # Fallback for old format
        items_dict = data
        total_time = 0
        
    # Sort keys
    sorted_keys = sorted(items_dict.keys(), key=lambda x: int(x))

    for key in sorted_keys:
        item = items_dict[key]
        
        jp_times = []
        jp_details = []
        
        if 'jump_points' in item and item['jump_points']:
            for jp in item['jump_points']:
                t_val = jp.get('time', 'N/A')
                jp_times.append(str(t_val))
                
                # Construct detail string: "(Tail, Xfade: 30)"
                type_str = jp.get('type', 'N/A')
                xfade_val = jp.get('xfade', 0)
                detail_str = f"({type_str}, Xfade: {xfade_val})"
                jp_details.append(detail_str)
        else:
            jp_times.append("None")
            jp_details.append("-")

        row = {
            "目标时长 (s)": item.get("target_duration"),
            "实际时长 (s)": item.get("actual_duration"),
            "处理耗时 (s)": item.get("process_time", 0), # New column
            "音频文件": item.get("files", {}).get("audio", ""),
            "跳点": "\n".join(jp_times),
            "(Tail, Xfade)": "\n".join(jp_details),
            "波形图": item.get("files", {}).get("image", "")
        }
        rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)
    
    cols = ["目标时长 (s)", "实际时长 (s)", "处理耗时 (s)", "音频文件", "跳点", "(Tail, Xfade)", "波形图"]
    df = df[cols]

    # Save to Excel
    print(f"DEBUG: Attempting to save to {output_path} with {len(df)} rows.")
    try:
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Use ExcelWriter to support multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Create Summary Sheet
            summary_data = [{"Project": "Batch Process Summary", "Total Time (s)": total_time, "Total Items": len(df)}]
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
        print(f"Successfully created Excel file at: {output_path}")
    except ImportError:
        print("Error: 'pandas' or 'openpyxl' library is missing. Please install them using: pip install pandas openpyxl")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def apply_conditional_formatting(excel_path):
    try:
        # Load workbook using openpyxl directly for formatting
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Define colors (hex codes)
        color_map = {
            2: "FFFF00", # Yellow
            3: "FFA500", # Orange
            4: "800080", # Purple
            5: "FF0000", # Red
            6: "A52A2A"  # Brown (and above)
        }
        
        # Helper to get lighter text color for dark backgrounds if needed, 
        # but for simplicity we'll keep black text for lighter colors and white for darker ones
        text_white = Font(color="FFFFFF")
        text_black = Font(color="000000")
        
        # Iterate over all columns
        for col in range(1, ws.max_column + 1):
            column_values = []
            
            # First pass: collect all values in this column (skip header)
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                # Convert to string to ensure consistent counting
                column_values.append(str(val) if val is not None else "")
                
            # Count frequencies
            from collections import Counter
            counts = Counter(column_values)
            
            # Second pass: apply formatting
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                val_str = str(cell.value) if cell.value is not None else ""
                
                freq = counts[val_str]
                
                fill_color = None
                font_style = text_black
                
                if freq == 2:
                    fill_color = color_map[2]
                elif freq == 3:
                    fill_color = color_map[3]
                elif freq == 4:
                    fill_color = color_map[4]
                    font_style = text_white # Purple is dark
                elif freq == 5:
                    fill_color = color_map[5]
                    font_style = text_white # Red is dark
                elif freq >= 6:
                    fill_color = color_map[6]
                    font_style = text_white # Brown is dark
                
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = font_style

        wb.save(excel_path)
        print(f"Applied conditional formatting to {excel_path}")
        
    except Exception as e:
        print(f"Error applying conditional formatting: {e}")

def insert_files_into_excel(excel_path, base_dir):
    try:
        import openpyxl
        from openpyxl.drawing.image import Image
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Find column indices
        headers = {cell.value: i for i, cell in enumerate(ws[1], 1)}
        img_col_idx = headers.get("波形图")
        audio_col_idx = headers.get("音频文件")
        
        if not img_col_idx:
            print("Could not find '波形图' column.")
            return

        # Set row height for images (approx 100 pixels height)
        # Excel row height is in points. 1 point ~= 1.33 pixels. 75 points ~= 100 px.
        # But we also have text. Let's make it taller.
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 80

        # Set column width (approx)
        ws.column_dimensions[openpyxl.utils.get_column_letter(img_col_idx)].width = 60
        
        for row in range(2, ws.max_row + 1):
            # 1. Insert Image
            img_filename = ws.cell(row=row, column=img_col_idx).value
            if img_filename:
                img_path = os.path.join(base_dir, str(img_filename))
                if os.path.exists(img_path):
                    try:
                        img = Image(img_path)
                        # Resize image to fit cell (approx 400x100 pixels)
                        # Keeping aspect ratio usually better, but we want it to fit the box
                        img.height = 100
                        img.width = 400
                        
                        # Calculate anchor, e.g. "D2"
                        col_letter = openpyxl.utils.get_column_letter(img_col_idx)
                        anchor = f"{col_letter}{row}"
                        
                        ws.add_image(img, anchor)
                        # Clear text content so it doesn't overlap/look weird
                        ws.cell(row=row, column=img_col_idx).value = "" 
                    except Exception as img_err:
                        print(f"Error inserting image {img_filename}: {img_err}")

            # 2. Insert Audio Link (Excel doesn't support embedding audio players easily in cells via openpyxl)
            # We can make it a hyperlink to the local file
            audio_filename = ws.cell(row=row, column=audio_col_idx).value
            if audio_filename:
                audio_path = os.path.join(base_dir, str(audio_filename))
                if os.path.exists(audio_path):
                    # Create hyperlink
                    # Absolute path is often safer for local links
                    abs_path = os.path.abspath(audio_path)
                    cell = ws.cell(row=row, column=audio_col_idx)
                    cell.hyperlink = abs_path
                    cell.value = audio_filename # Keep the name visible
                    cell.style = "Hyperlink"

        wb.save(excel_path)
        print(f"Inserted images and links into {excel_path}")

    except Exception as e:
        print(f"Error inserting files: {e}")

if __name__ == "__main__":
    json_file = r"batch_output\batch_info.json"
    excel_file = r"batch_output\batch_info_media.xlsx" # New filename
    output_dir = r"batch_output"
    
    print(f"Converting {json_file} to {excel_file}...")
    json_to_excel(json_file, excel_file)
    apply_conditional_formatting(excel_file)
    insert_files_into_excel(excel_file, output_dir)
